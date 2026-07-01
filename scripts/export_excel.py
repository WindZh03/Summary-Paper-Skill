#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

from paper_summary_lib import EXCEL_COLUMN_MAP, first_nonempty, load_summary_rows


def import_openpyxl():
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Alignment, Font  # type: ignore

        return Workbook, Alignment, Font
    except ImportError as exc:
        raise SystemExit(
            "Missing Excel dependency. Run 'python3 scripts/check_deps.py check' first."
        ) from exc


def excel_column_name(index: int) -> str:
    letters = []
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def export_excel(summary_json: Path, output: Path) -> None:
    Workbook, Alignment, Font = import_openpyxl()
    rows = load_summary_rows(summary_json)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "papers"
    sheet.freeze_panes = "A2"

    for col_idx, (header, _) in enumerate(EXCEL_COLUMN_MAP, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_, keys) in enumerate(EXCEL_COLUMN_MAP, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=first_nonempty(row, *keys))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "源文件": 28,
        "源路径": 56,
        "title": 42,
        "publish+time": 18,
        "keywords": 28,
        "研究现状": 36,
        "motivation": 36,
        "insight": 36,
        "核心贡献": 48,
        "method": 72,
        "实验结论": 48,
        "局限性": 48,
        "其它": 48,
    }
    for col_idx, (header, _) in enumerate(EXCEL_COLUMN_MAP, start=1):
        sheet.column_dimensions[excel_column_name(col_idx)].width = widths[header]

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Export structured paper summaries to one XLSX sheet.")
    parser.add_argument("summary_json")
    parser.add_argument("--output", required=True)
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    output = Path(args.output).expanduser().resolve()
    export_excel(Path(args.summary_json).expanduser().resolve(), output)
    print(str(output))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
